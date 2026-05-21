"""Styles Excel centralisés pour l'export."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Couleurs
NAVY = "1F3864"          # titre principal
LIGHT_BLUE = "D9E2F3"    # sections / headers
GREY_TOTAL = "F2F2F2"    # fond total
GREY_RATIO = "7F7F7F"    # texte ratio italique

# Polices
FONT_TITLE = Font(name="Arial", size=12, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Arial", size=9, italic=True, color="7F7F7F")
FONT_HEADER = Font(name="Arial", size=10, bold=True)
FONT_SECTION = Font(name="Arial", size=10, bold=True)
FONT_VALUE = Font(name="Arial", size=10)
FONT_RATIO = Font(name="Arial", size=10, italic=True, color=GREY_RATIO)
FONT_TOTAL = Font(name="Arial", size=10, bold=True)
FONT_CHECK_OK = Font(name="Arial", size=9, color="008000")
FONT_CHECK_WARN = Font(name="Arial", size=9, color="C00000")
FONT_CHECK_NEUTRAL = Font(name="Arial", size=9, color="7F7F7F")

# Remplissages
FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_SECTION = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_TOTAL = PatternFill("solid", fgColor=GREY_TOTAL)

# Alignements
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

# Bordures
_thin = Side(style="thin", color="000000")
BORDER_TOTAL_TOP = Border(top=_thin)


def get_money_format(currency: str) -> str:
    """Retourne un format Excel adapté à la devise."""
    symbol_map = {"USD": "$", "EUR": "€", "GBP": "£", "JPY": "¥", "CHF": "CHF "}
    symbol = symbol_map.get(currency.upper(), "")
    if symbol:
        return f'_({symbol}* # ##0_);_({symbol}* (# ##0);_({symbol}* "-"_);_(@_)'
    # Fallback avec code devise en suffixe
    return f'_(* # ##0_ "{currency}");_(* (# ##0)" {currency}";_(* "-"_ "{currency}");_(@_)'


PCT_FORMAT = '0.0%;(0.0%);"-"'
RATIO_FORMAT = '0.00"x";(0.00)"x";"-"'