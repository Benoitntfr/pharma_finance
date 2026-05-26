"""Export Excel des 5 blocs, mis en forme."""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from src.excel_styles import (
    FONT_TITLE, FONT_SUBTITLE, FONT_HEADER, FONT_SECTION, FONT_VALUE,
    FONT_RATIO, FONT_TOTAL,
    FONT_CHECK_OK, FONT_CHECK_WARN, FONT_CHECK_NEUTRAL,
    FILL_TITLE, FILL_SECTION, FILL_TOTAL,
    ALIGN_LEFT, ALIGN_RIGHT, ALIGN_CENTER,
    BORDER_TOTAL_TOP,
    get_money_format, PCT_FORMAT, RATIO_FORMAT,
)
from src.utils import safe_div


# ============================================================
# Scale adaptatif : k ou M selon la taille de la boîte
# ============================================================

def _decide_scale(pl_data: dict):
    """
    Décide l'échelle selon le revenue le plus récent.
    Retourne (divisor, label).
    Seuil : 10B revenue → M, sinon → k.
    """
    rev_last = pl_data["rows"]["Revenue"][2]
    if rev_last and abs(rev_last) >= 10e9:
        return 1_000_000, "millions"
    return 1_000, "thousands"


def _apply_scale(values: list, divisor: int) -> list:
    """Divise une liste de valeurs par le divisor, en préservant les None."""
    return [v / divisor if v is not None else None for v in values]


# ============================================================
# Entry point
# ============================================================

def export_to_excel(
    ticker: str,
    info: dict,
    identity: dict,
    pl_data: dict,
    bs_data: dict,
    cf_data: dict,
    ratios: dict,
    currency: str,
) -> str:
    """Génère un Excel formaté avec 5 onglets. Retourne le chemin."""
    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{ticker}_{date_str}.xlsx"
    os.makedirs("exports", exist_ok=True)
    filepath = os.path.join("exports", filename)

    scale_div, scale_label = _decide_scale(pl_data)

    wb = Workbook()
    wb.remove(wb.active)

    company_name = identity.get("name", "")
    exchange = info.get("exchange", "")
    money_fmt = get_money_format(currency)

    _create_identity_sheet(wb, identity, company_name, exchange, currency)
    _create_pl_sheet(wb, pl_data, company_name, exchange, ticker, currency, money_fmt, scale_div, scale_label)
    _create_bs_sheet(wb, bs_data, company_name, exchange, ticker, currency, money_fmt, scale_div, scale_label)
    _create_cf_sheet(wb, cf_data, company_name, exchange, ticker, currency, money_fmt, scale_div, scale_label)
    _create_ratios_sheet(wb, ratios, company_name, exchange, ticker, currency, money_fmt)

    wb.save(filepath)
    return filepath


# ============================================================
# Helpers communs
# ============================================================

def _setup_sheet(ws, title_text: str, subtitle_text: str, num_cols: int):
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    cell.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    cell = ws.cell(row=2, column=1, value=subtitle_text)
    cell.font = FONT_SUBTITLE
    cell.alignment = ALIGN_LEFT

    ws.row_dimensions[3].height = 8


def _write_section_header(ws, row: int, label: str, num_cols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = ALIGN_LEFT


def _write_col_headers(ws, row: int, headers: list):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_SECTION
        cell.alignment = ALIGN_RIGHT if i > 0 else ALIGN_LEFT


def _write_value_row(ws, row: int, label: str, values: list, num_fmt: str,
                     is_total: bool = False, is_ratio: bool = False):
    label_cell = ws.cell(row=row, column=1, value=label)
    if is_total:
        label_cell.font = FONT_TOTAL
        label_cell.fill = FILL_TOTAL
        label_cell.border = BORDER_TOTAL_TOP
    elif is_ratio:
        label_cell.font = FONT_RATIO
    else:
        label_cell.font = FONT_VALUE
    label_cell.alignment = ALIGN_LEFT

    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=i + 2, value=v)
        cell.number_format = num_fmt
        cell.alignment = ALIGN_RIGHT
        if is_total:
            cell.font = FONT_TOTAL
            cell.fill = FILL_TOTAL
            cell.border = BORDER_TOTAL_TOP
        elif is_ratio:
            cell.font = FONT_RATIO
        else:
            cell.font = FONT_VALUE


def _set_col_widths(ws, label_width: int = 38, value_width: int = 16, num_value_cols: int = 3):
    ws.column_dimensions["A"].width = label_width
    for i in range(num_value_cols):
        ws.column_dimensions[get_column_letter(i + 2)].width = value_width


def _write_checks(ws, start_row: int, checks: list, num_cols: int) -> int:
    start_row += 2
    _write_section_header(ws, start_row, "Checks de cohérence", num_cols)
    start_row += 1
    for status, text in checks:
        cell = ws.cell(row=start_row, column=1, value=text)
        if status == "ok":
            cell.font = FONT_CHECK_OK
        elif status == "warn":
            cell.font = FONT_CHECK_WARN
        else:
            cell.font = FONT_CHECK_NEUTRAL
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
        start_row += 1
    return start_row


# ============================================================
# Onglets
# ============================================================

def _create_identity_sheet(wb, identity, company_name, exchange, currency):
    ws = wb.create_sheet("Identity")
    title = f"{company_name} ({exchange}: {identity.get('ticker', '')}) - Identity"
    _setup_sheet(ws, title, f"All figures in {currency}.", 2)
    _set_col_widths(ws, label_width=30, value_width=40, num_value_cols=1)

    def fmt_employees(n):
        return f"{n:,}" if n else "N/A"

    def fmt_money_raw(v):
        if v is None:
            return "N/A"
        if abs(v) >= 1e9:
            return f"{v/1e9:.2f}B {currency}"
        if abs(v) >= 1e6:
            return f"{v/1e6:.1f}M {currency}"
        return f"{v:,.0f} {currency}"

    rows = [
        ("Name", identity.get("name") or "N/A"),
        ("Ticker", identity.get("ticker") or "N/A"),
        ("Market Cap", fmt_money_raw(identity.get("market_cap"))),
        ("Enterprise Value", fmt_money_raw(identity.get("enterprise_value"))),
        ("Country", identity.get("country") or "N/A"),
        ("Sector", identity.get("sector") or "N/A"),
        ("Industry", identity.get("industry") or "N/A"),
        ("Employees", fmt_employees(identity.get("employees"))),
        ("Currency (market)", identity.get("currency") or "N/A"),
        ("Currency (financials)", identity.get("financial_currency") or "N/A"),
    ]

    row = 4
    _write_col_headers(ws, row, ["Field", "Value"])
    row += 1
    for label, value in rows:
        ws.cell(row=row, column=1, value=label).font = FONT_VALUE
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=value).font = FONT_VALUE
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT
        row += 1

    summary = identity.get("summary")
    if summary:
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value="Business summary")
        cell.font = FONT_SECTION
        cell.fill = FILL_SECTION
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=summary)
        cell.font = FONT_SUBTITLE
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 80


def _create_pl_sheet(wb, pl_data, company_name, exchange, ticker, currency, money_fmt, scale_div, scale_label):
    ws = wb.create_sheet("P&L")
    years = pl_data["years"]
    rows = pl_data["rows"]
    title = f"{company_name} ({exchange}: {ticker}) - Profit & Loss"
    _setup_sheet(ws, title, f"All figures in {currency} {scale_label} unless noted.", 4)
    _set_col_widths(ws, label_width=38, value_width=16, num_value_cols=3)

    rev = _apply_scale(rows["Revenue"], scale_div)
    gp = _apply_scale(rows["Gross Profit"], scale_div)
    rd = _apply_scale(rows["R&D"], scale_div)
    sga = _apply_scale(rows["SG&A"], scale_div)
    ebit = _apply_scale(rows["EBIT"], scale_div)
    ebitda = _apply_scale(rows["EBITDA"], scale_div)
    ni = _apply_scale(rows["Net Income"], scale_div)

    yoy = [
        None,
        (rev[1] / rev[0] - 1) if (rev[0] and rev[1]) else None,
        (rev[2] / rev[1] - 1) if (rev[1] and rev[2]) else None,
    ]
    gm = [safe_div(gp[i], rev[i]) for i in range(3)]
    rd_pct = [safe_div(rd[i], rev[i]) for i in range(3)]
    sga_pct = [safe_div(sga[i], rev[i]) for i in range(3)]
    ebit_m = [safe_div(ebit[i], rev[i]) for i in range(3)]
    ebitda_m = [safe_div(ebitda[i], rev[i]) for i in range(3)]
    ni_m = [safe_div(ni[i], rev[i]) for i in range(3)]

    has_ebitda = any(v is not None for v in ebitda)

    row = 4
    _write_col_headers(ws, row, ["", f"FY{years[0]}", f"FY{years[1]}", f"FY{years[2]}"])
    row += 1

    _write_section_header(ws, row, "Revenue", 4); row += 1
    _write_value_row(ws, row, "Total Revenue", rev, money_fmt, is_total=True); row += 1
    _write_value_row(ws, row, "  YoY %", yoy, PCT_FORMAT, is_ratio=True); row += 1

    _write_section_header(ws, row, "Profitability", 4); row += 1
    _write_value_row(ws, row, "Gross Profit", gp, money_fmt); row += 1
    _write_value_row(ws, row, "  Gross Margin", gm, PCT_FORMAT, is_ratio=True); row += 1
    _write_value_row(ws, row, "EBIT", ebit, money_fmt); row += 1
    _write_value_row(ws, row, "  EBIT Margin", ebit_m, PCT_FORMAT, is_ratio=True); row += 1
    if has_ebitda:
        _write_value_row(ws, row, "EBITDA", ebitda, money_fmt); row += 1
        _write_value_row(ws, row, "  EBITDA Margin", ebitda_m, PCT_FORMAT, is_ratio=True); row += 1
    _write_value_row(ws, row, "Net Income", ni, money_fmt, is_total=True); row += 1
    _write_value_row(ws, row, "  Net Margin", ni_m, PCT_FORMAT, is_ratio=True); row += 1

    _write_section_header(ws, row, "Operating expenses", 4); row += 1
    _write_value_row(ws, row, "R&D", rd, money_fmt); row += 1
    _write_value_row(ws, row, "  % Revenue", rd_pct, PCT_FORMAT, is_ratio=True); row += 1
    _write_value_row(ws, row, "SG&A", sga, money_fmt); row += 1
    _write_value_row(ws, row, "  % Revenue", sga_pct, PCT_FORMAT, is_ratio=True); row += 1


def _create_bs_sheet(wb, bs_data, company_name, exchange, ticker, currency, money_fmt, scale_div, scale_label):
    ws = wb.create_sheet("Balance Sheet")
    years = bs_data["years"]
    rows = bs_data["rows"]
    title = f"{company_name} ({exchange}: {ticker}) - Balance Sheet"
    _setup_sheet(ws, title, f"All figures in {currency} {scale_label}.", 3)
    _set_col_widths(ws, label_width=38, value_width=18, num_value_cols=2)

    ta = _apply_scale(rows["Total Assets"], scale_div)
    ca = _apply_scale(rows["Current Assets"], scale_div)
    nca = _apply_scale(rows["Non-Current Assets"], scale_div)
    tl = _apply_scale(rows["Total Liabilities"], scale_div)
    cl = _apply_scale(rows["Current Liabilities"], scale_div)
    ncl = _apply_scale(rows["Non-Current Liabilities"], scale_div)
    eq = _apply_scale(rows["Total Equity"], scale_div)
    cash = _apply_scale(rows["Cash & ST Investments"], scale_div)
    cdebt = _apply_scale(rows["Current Debt"], scale_div)
    ltdebt = _apply_scale(rows["Long-Term Debt"], scale_div)
    debt = _apply_scale(rows["Total Debt"], scale_div)
    gw = _apply_scale(rows["Goodwill"], scale_div)
    intang = _apply_scale(rows["Intangible Assets"], scale_div)

    net_debt = [
        (debt[i] - cash[i]) if (debt[i] is not None and cash[i] is not None) else None
        for i in range(2)
    ]
    gw_ratio = [safe_div(gw[i], ta[i]) for i in range(2)]

    row = 4
    _write_col_headers(ws, row, ["", f"FY{years[0]}", f"FY{years[1]}"])
    row += 1

    _write_section_header(ws, row, "Assets", 3); row += 1
    _write_value_row(ws, row, "  Current Assets", ca, money_fmt); row += 1
    _write_value_row(ws, row, "  Non-Current Assets", nca, money_fmt); row += 1
    _write_value_row(ws, row, "Total Assets", ta, money_fmt, is_total=True); row += 1

    _write_section_header(ws, row, "Liabilities & Equity", 3); row += 1
    _write_value_row(ws, row, "  Current Liabilities", cl, money_fmt); row += 1
    _write_value_row(ws, row, "  Non-Current Liabilities", ncl, money_fmt); row += 1
    _write_value_row(ws, row, "Total Liabilities", tl, money_fmt, is_total=True); row += 1
    _write_value_row(ws, row, "Total Equity", eq, money_fmt, is_total=True); row += 1

    _write_section_header(ws, row, "Debt & Cash", 3); row += 1
    _write_value_row(ws, row, "Cash & ST Investments", cash, money_fmt); row += 1
    _write_value_row(ws, row, "Current Debt", cdebt, money_fmt); row += 1
    _write_value_row(ws, row, "Long-Term Debt", ltdebt, money_fmt); row += 1
    _write_value_row(ws, row, "Total Debt", debt, money_fmt, is_total=True); row += 1
    _write_value_row(ws, row, "Net Debt", net_debt, money_fmt); row += 1

    _write_section_header(ws, row, "Intangibles", 3); row += 1
    _write_value_row(ws, row, "Goodwill", gw, money_fmt); row += 1
    _write_value_row(ws, row, "  GW / Total Assets", gw_ratio, PCT_FORMAT, is_ratio=True); row += 1
    _write_value_row(ws, row, "Intangible Assets", intang, money_fmt); row += 1

    checks = _build_bs_checks(ta[1], tl[1], eq[1], gw[1])
    _write_checks(ws, row, checks, 3)


def _build_bs_checks(ta, tl, eq, gw):
    checks = []
    if ta is not None and tl is not None and eq is not None:
        sum_le = tl + eq
        diff_pct = abs(ta - sum_le) / ta if ta != 0 else 0
        if diff_pct < 0.01:
            checks.append(("ok", f"✓ Total Assets = L + E (écart {diff_pct*100:.2f}%)"))
        else:
            checks.append(("warn", f"✗ Total Assets ≠ L + E (écart {diff_pct*100:.2f}%)"))
    if gw is not None and ta is not None and ta != 0:
        ratio = gw / ta
        if ratio > 0.40:
            checks.append(("warn", f"⚠ Goodwill / Total Assets = {ratio*100:.1f}% > 40%"))
        else:
            checks.append(("ok", f"✓ Goodwill / Total Assets = {ratio*100:.1f}% (sous 40%)"))
    return checks

def _create_cf_sheet(wb, cf_data, company_name, exchange, ticker, currency, money_fmt, scale_div, scale_label):
    ws = wb.create_sheet("Cash Flow")
    years = cf_data["years"]
    rows = cf_data["rows"]
    title = f"{company_name} ({exchange}: {ticker}) - Cash Flow"
    _setup_sheet(ws, title, f"All figures in {currency} {scale_label}.", 4)
    _set_col_widths(ws, label_width=42, value_width=16, num_value_cols=3)

    # Apply scale to all rows
    def s(key):
        return _apply_scale(rows[key], scale_div)

    ni = s("Net Income")
    da = s("D&A")
    sbc = s("Stock-Based Compensation")
    def_tax = s("Deferred Taxes")
    other_nc = s("Other Non-Cash")
    d_rec = s("Δ Receivables")
    d_inv = s("Δ Inventory")
    d_pay = s("Δ Payables")
    d_other_wc = s("Δ Other WC")
    total_d_wc = s("Total Δ WC")
    cfo = s("CFO")

    capex = s("CapEx")
    acq = s("Acquisitions")
    div_st = s("Divestitures")
    net_ma = s("Net M&A")
    pinv = s("Purchase Of Investments")
    sinv = s("Sale Of Investments")
    net_inv = s("Net Investments")
    other_inv = s("Other Investing")
    cfi = s("CFI")

    debt_iss = s("Debt Issued")
    debt_rep = s("Debt Repaid")
    net_debt = s("Net Debt Change")
    eq_iss = s("Equity Issued")
    bb = s("Buybacks")
    net_cs = s("Net Common Stock Issuance")
    div = s("Dividends Paid")
    other_fin = s("Other Financing")
    cff = s("CFF")

    fx = s("FX Effect")
    net_change = s("Net Change In Cash")
    beg_cash = s("Beginning Cash")
    end_cash = s("Ending Cash")
    fcf = s("FCF")

    fcf_calc = [
        (cfo[i] - abs(capex[i])) if (cfo[i] is not None and capex[i] is not None) else None
        for i in range(3)
    ]

    row = 4
    _write_col_headers(ws, row, ["", f"FY{years[0]}", f"FY{years[1]}", f"FY{years[2]}"])
    row += 1

    # Net Income (start)
    _write_value_row(ws, row, "Net Income", ni, money_fmt, is_total=True); row += 1

    _write_section_header(ws, row, "Non-cash adjustments", 4); row += 1
    _write_value_row(ws, row, "  D&A", da, money_fmt); row += 1
    _write_value_row(ws, row, "  Stock-Based Compensation", sbc, money_fmt); row += 1
    _write_value_row(ws, row, "  Deferred Taxes", def_tax, money_fmt); row += 1
    _write_value_row(ws, row, "  Other Non-Cash", other_nc, money_fmt); row += 1

    _write_section_header(ws, row, "Working Capital", 4); row += 1
    _write_value_row(ws, row, "  Δ Receivables", d_rec, money_fmt); row += 1
    _write_value_row(ws, row, "  Δ Inventory", d_inv, money_fmt); row += 1
    _write_value_row(ws, row, "  Δ Payables", d_pay, money_fmt); row += 1
    _write_value_row(ws, row, "  Δ Other WC", d_other_wc, money_fmt); row += 1
    _write_value_row(ws, row, "  Total Δ WC", total_d_wc, money_fmt); row += 1

    _write_value_row(ws, row, "CFO (Operating Cash Flow)", cfo, money_fmt, is_total=True); row += 1

    _write_section_header(ws, row, "Capital expenditure", 4); row += 1
    _write_value_row(ws, row, "  CapEx", capex, money_fmt); row += 1

    _write_section_header(ws, row, "Acquisitions & divestitures", 4); row += 1
    _write_value_row(ws, row, "  Acquisitions", acq, money_fmt); row += 1
    _write_value_row(ws, row, "  Divestitures", div_st, money_fmt); row += 1
    _write_value_row(ws, row, "  Net M&A", net_ma, money_fmt); row += 1

    _write_section_header(ws, row, "Investment activities", 4); row += 1
    _write_value_row(ws, row, "  Purchase of investments", pinv, money_fmt); row += 1
    _write_value_row(ws, row, "  Sale/maturity of investments", sinv, money_fmt); row += 1
    _write_value_row(ws, row, "  Net investments", net_inv, money_fmt); row += 1
    _write_value_row(ws, row, "  Other investing", other_inv, money_fmt); row += 1

    _write_value_row(ws, row, "CFI (Investing Cash Flow)", cfi, money_fmt, is_total=True); row += 1

    _write_section_header(ws, row, "Debt activity", 4); row += 1
    _write_value_row(ws, row, "  Debt issued", debt_iss, money_fmt); row += 1
    _write_value_row(ws, row, "  Debt repaid", debt_rep, money_fmt); row += 1
    _write_value_row(ws, row, "  Net debt change", net_debt, money_fmt); row += 1

    _write_section_header(ws, row, "Equity activity", 4); row += 1
    _write_value_row(ws, row, "  Equity issued", eq_iss, money_fmt); row += 1
    _write_value_row(ws, row, "  Buybacks", bb, money_fmt); row += 1
    _write_value_row(ws, row, "  Net common stock", net_cs, money_fmt); row += 1

    _write_section_header(ws, row, "Distributions", 4); row += 1
    _write_value_row(ws, row, "  Dividends paid", div, money_fmt); row += 1
    _write_value_row(ws, row, "  Other financing", other_fin, money_fmt); row += 1

    _write_value_row(ws, row, "CFF (Financing Cash Flow)", cff, money_fmt, is_total=True); row += 1

    _write_section_header(ws, row, "FX & Net change", 4); row += 1
    _write_value_row(ws, row, "  FX effect", fx, money_fmt); row += 1
    _write_value_row(ws, row, "Net change in cash", net_change, money_fmt, is_total=True); row += 1

    _write_section_header(ws, row, "Reconciliation", 4); row += 1
    _write_value_row(ws, row, "  Beginning cash", beg_cash, money_fmt); row += 1
    _write_value_row(ws, row, "  Ending cash", end_cash, money_fmt); row += 1

    _write_section_header(ws, row, "Free Cash Flow", 4); row += 1
    _write_value_row(ws, row, "  FCF (yfinance)", fcf, money_fmt); row += 1
    _write_value_row(ws, row, "  FCF reconstruit (CFO - |CapEx|)", fcf_calc, money_fmt); row += 1

    # Checks
    checks = _build_cf_checks_extended(cf_data, fcf, fcf_calc, scale_div)
    _write_checks(ws, row, checks, 4)


def _build_cf_checks_extended(cf_data, fcf_scaled, fcf_calc_scaled, scale_div):
    """Checks Cash Flow étendus : CFO articulation, FCF, Cash reconciliation."""
    rows = cf_data["rows"]
    # On utilise les valeurs originales (non scalées) pour les ratios
    ni = rows["Net Income"][2]
    da = rows["D&A"][2]
    sbc = rows["Stock-Based Compensation"][2]
    def_tax = rows["Deferred Taxes"][2]
    other_nc = rows["Other Non-Cash"][2]
    total_d_wc = rows["Total Δ WC"][2]
    cfo_yf = rows["CFO"][2]
    fcf_yf = rows["FCF"][2]
    cfi_n = rows["CFI"][2]
    cff_n = rows["CFF"][2]
    fx_n = rows["FX Effect"][2]
    beg_cash = rows["Beginning Cash"][2]
    end_cash = rows["Ending Cash"][2]

    checks = []

    # Check 1 : CFO articulation
    components = [ni, da, sbc, def_tax, other_nc, total_d_wc]
    known = [v for v in components if v is not None]
    if cfo_yf is not None and len(known) >= 3:
        sum_known = sum(known)
        diff_pct = abs(cfo_yf - sum_known) / abs(cfo_yf) if cfo_yf != 0 else 0
        if diff_pct < 0.05:
            checks.append(("ok", f"✓ CFO articulation : NI + ajustements ≈ CFO (écart {diff_pct*100:.1f}%)"))
        else:
            checks.append(("warn", f"⚠ CFO articulation : écart {diff_pct*100:.1f}% (lignes 'Other' non détaillées)"))
    else:
        checks.append(("neutral", "⊘ CFO articulation : données insuffisantes"))

    # Check 2 : FCF reconstruction
    if fcf_yf is not None and fcf_calc_scaled[2] is not None:
        fcf_calc_n = fcf_calc_scaled[2] * scale_div  # re-unscale pour comparaison
        diff_pct = abs(fcf_yf - fcf_calc_n) / abs(fcf_yf) if fcf_yf != 0 else 0
        if diff_pct < 0.01:
            checks.append(("ok", f"✓ FCF yfinance ≈ FCF reconstruit (écart {diff_pct*100:.2f}%)"))
        else:
            checks.append(("warn", f"⚠ FCF yfinance vs reconstruit : écart {diff_pct*100:.2f}%"))
    else:
        checks.append(("neutral", "⊘ FCF check : données manquantes"))

    # Check 3 : Cash reconciliation
    if (beg_cash is not None and end_cash is not None
            and cfo_yf is not None and cfi_n is not None and cff_n is not None):
        fx_val = fx_n if fx_n is not None else 0
        calc_end = beg_cash + cfo_yf + cfi_n + cff_n + fx_val
        diff_pct = abs(calc_end - end_cash) / abs(end_cash) if end_cash != 0 else 0
        if diff_pct < 0.01:
            checks.append(("ok", f"✓ Cash reconciliation : Beg + CFO + CFI + CFF + FX ≈ Ending Cash (écart {diff_pct*100:.2f}%)"))
        else:
            checks.append(("warn", f"⚠ Cash reconciliation : écart {diff_pct*100:.2f}%"))
    else:
        checks.append(("neutral", "⊘ Cash reconciliation : Beginning/Ending Cash absents"))

    return checks

def _create_ratios_sheet(wb, ratios, company_name, exchange, ticker, currency, money_fmt):
    ws = wb.create_sheet("Ratios")
    title = f"{company_name} ({exchange}: {ticker}) - Ratios"
    _setup_sheet(ws, title, "Derived ratios and cross-bloc checks.", 2)
    _set_col_widths(ws, label_width=42, value_width=22, num_value_cols=1)

    years = ratios["years"]

    def fmt_pct(v):
        return f"{v*100:.1f}%" if v is not None else "N/A"

    def fmt_ratio(v, nm=False):
        if nm:
            return "N/M"
        return f"{v:.2f}x" if v is not None else "N/A"

    def fmt_money(v):
        if v is None:
            return "N/A"
        if abs(v) >= 1e9:
            return f"{v/1e9:.2f}B {currency}"
        if abs(v) >= 1e6:
            return f"{v/1e6:.1f}M {currency}"
        return f"{v:,.0f} {currency}"

    row = 4
    _write_col_headers(ws, row, ["Ratio / Indicator", "Value"])
    row += 1

    _write_section_header(ws, row, "Growth & profitability", 2); row += 1
    pairs = [
        ("Revenue CAGR 3Y", fmt_pct(ratios["cagr"])),
        (f"R&D / Revenue ({years[0]} → {years[2]})",
         " → ".join(fmt_pct(v) for v in ratios["rd_trend"])),
        (f"EBIT margin ({years[0]} → {years[2]})",
         " → ".join(fmt_pct(v) for v in ratios["ebit_trend"])),
    ]
    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = FONT_VALUE
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=value).font = FONT_VALUE
        ws.cell(row=row, column=2).alignment = ALIGN_RIGHT
        row += 1

    _write_section_header(ws, row, "Capital structure & cash generation", 2); row += 1
    roic_str = "N/M" if ratios["roic_nm"] else fmt_pct(ratios["roic"])
    pairs = [
        ("FCF / Net Income (N)", fmt_ratio(ratios["fcf_conv"], ratios["fcf_conv_nm"])),
        ("Net Debt / EBITDA (N)", fmt_ratio(ratios["nd_ebitda"], ratios["nd_ebitda_nm"])),
        ("Goodwill / Total Assets (N)", fmt_pct(ratios["gw_ta"])),
        (f"ROIC (N) — tax {ratios['tax_rate_source']}", roic_str),
    ]
    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = FONT_VALUE
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=value).font = FONT_VALUE
        ws.cell(row=row, column=2).alignment = ALIGN_RIGHT
        row += 1

    _write_section_header(ws, row, "EV reconciliation", 2); row += 1
    pairs = [
        ("EV reconstruit (Mcap + Debt - Cash)", fmt_money(ratios["ev_calc"])),
        ("EV yfinance", fmt_money(ratios["ev_info"])),
        ("EV/Revenue calculé",
         f"{ratios['ev_rev_calc']:.2f}x" if ratios["ev_rev_calc"] else "N/A"),
        ("EV/Revenue yfinance",
         f"{ratios['ev_rev_info']:.2f}x" if ratios["ev_rev_info"] else "N/A"),
        ("EV/EBITDA calculé",
         f"{ratios['ev_ebitda_calc']:.2f}x" if ratios["ev_ebitda_calc"] else "N/A"),
        ("EV/EBITDA yfinance",
         f"{ratios['ev_ebitda_info']:.2f}x" if ratios["ev_ebitda_info"] else "N/A"),
    ]
    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = FONT_VALUE
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=value).font = FONT_VALUE
        ws.cell(row=row, column=2).alignment = ALIGN_RIGHT
        row += 1

    checks = _build_ratio_checks(ratios)
    _write_checks(ws, row, checks, 2)


def _build_ratio_checks(r):
    checks = []
    if r["ev_calc"] is not None and r["ev_info"] is not None and r["ev_info"] != 0:
        d = abs(r["ev_calc"] - r["ev_info"]) / r["ev_info"]
        status = "ok" if d < 0.05 else "warn"
        sym = "✓" if d < 0.05 else "⚠"
        checks.append((status, f"{sym} EV reconstruit vs yfinance : écart {d*100:.1f}%"))
    if r["ev_rev_calc"] is not None and r["ev_rev_info"] is not None and r["ev_rev_info"] != 0:
        d = abs(r["ev_rev_calc"] - r["ev_rev_info"]) / r["ev_rev_info"]
        status = "ok" if d < 0.05 else "warn"
        sym = "✓" if d < 0.05 else "⚠"
        checks.append((status, f"{sym} EV/Revenue calculé vs yfinance : écart {d*100:.1f}%"))
    if r["ev_ebitda_calc"] is not None and r["ev_ebitda_info"] is not None and r["ev_ebitda_info"] != 0:
        d = abs(r["ev_ebitda_calc"] - r["ev_ebitda_info"]) / r["ev_ebitda_info"]
        status = "ok" if d < 0.05 else "warn"
        sym = "✓" if d < 0.05 else "⚠"
        checks.append((status, f"{sym} EV/EBITDA calculé vs yfinance : écart {d*100:.1f}%"))
    else:
        checks.append(("neutral", "⊘ EV/EBITDA : EBITDA négatif ou données manquantes"))
    return checks


def download_if_colab(filepath: str) -> None:
    """Sur Colab, déclenche le téléchargement automatique."""
    try:
        from google.colab import files
        files.download(filepath)
    except ImportError:
        print(f"📁 Fichier exporté : {filepath}")